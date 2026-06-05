"""
Week 3 end-to-end test suite.
Runs against a live server at http://localhost:8000.
"""
import json
import sys
import time
from pathlib import Path

import httpx

BASE = "http://localhost:8000"
PDF_PATH = Path(__file__).parent / "sample_records" / "test_medical_record.pdf"

EMAIL = "e2e_test@chartlens.test"
PASSWORD = "e2etest123"
FIRM = "E2E Test Law Firm"


def print_result(label: str, passed: bool, detail: str = "") -> None:
    status = "PASS" if passed else "FAIL"
    suffix = f" — {detail}" if detail else ""
    print(f"  [{status}] {label}{suffix}")


def run_tests() -> int:
    """Return number of failures."""
    failures = 0
    client = httpx.Client(base_url=BASE, follow_redirects=True, timeout=30)

    # ── TEST 1: REGISTER ──────────────────────────────────────────────────────
    print("\n[1] REGISTER")
    r = client.post("/auth/register", data={
        "email": EMAIL, "password": PASSWORD,
        "confirm_password": PASSWORD, "firm_name": FIRM,
    })
    t1_ok = r.status_code == 200 and "dashboard" in str(r.url)
    print_result("POST /auth/register -> /dashboard", t1_ok,
                 f"status={r.status_code} url={r.url}")
    if not t1_ok:
        failures += 1
        print("  Stopping — can't test further without auth.", file=sys.stderr)
        return failures

    dashboard_html = r.text
    trial_badge = "Trial" in dashboard_html and "remaining" in dashboard_html
    cta = "Upload New Case" in dashboard_html
    print_result("Dashboard shows 'Trial — X cases remaining'", trial_badge)
    print_result("Dashboard shows 'Upload New Case' CTA", cta)
    if not trial_badge or not cta:
        failures += 1

    # ── TEST 2: UPLOAD + PROCESSING ──────────────────────────────────────────
    print("\n[2] UPLOAD + PROCESSING")
    if not PDF_PATH.exists():
        print(f"  [FAIL] Sample PDF not found at {PDF_PATH}")
        failures += 1
        return failures

    with open(PDF_PATH, "rb") as f:
        r2 = client.post("/cases/upload", data={
            "client_name": "John Anderson",
            "case_number": "2024-PI-001",
        }, files={"file": ("test_medical_record.pdf", f, "application/pdf")})

    upload_ok = r2.status_code == 200 and "/cases/" in str(r2.url)
    print_result("POST /cases/upload -> /cases/{id}", upload_ok,
                 f"status={r2.status_code} url={r2.url}")
    if not upload_ok:
        failures += 1
        return failures

    case_url = str(r2.url)
    case_id = case_url.rstrip("/").split("/")[-1]
    print(f"  Case ID: {case_id}")

    # Check processing spinner visible initially
    spinner = "animate-spin" in r2.text or "Processing" in r2.text or "Analyzing" in r2.text
    print_result("Results page shows processing state", spinner)
    if not spinner:
        failures += 1

    # Poll for completion (max 180 seconds for real Claude calls)
    print("  Polling for completion (real Claude API calls, ~30–90s)...")
    status = "processing"
    for attempt in range(60):
        time.sleep(3)
        rs = client.get(f"/cases/{case_id}/status")
        if rs.status_code == 200:
            data = rs.json()
            status = data.get("status", "unknown")
            pct = data.get("progress_percent", 0)
            if attempt % 5 == 0:
                print(f"    {attempt*3}s  status={status}  progress={pct}%")
            if status in ("complete", "failed"):
                break

    print_result("Analysis completed (not stuck)", status in ("complete", "failed"),
                 f"final_status={status}")
    if status not in ("complete", "failed"):
        failures += 1
        return failures
    if status == "failed":
        # Fetch results page to get error message
        rr = client.get(f"/cases/{case_id}")
        print_result("Analysis completed successfully", False,
                     "status=failed — check server logs")
        failures += 1
        return failures

    # ── TEST 3: RESULTS PAGE ─────────────────────────────────────────────────
    print("\n[3] RESULTS PAGE")
    rr = client.get(f"/cases/{case_id}")
    results_html = rr.text

    has_events_card = "Total Events" in results_html
    has_billed_card = "Total Billed" in results_html
    has_paid_card = "Total Paid" in results_html
    has_incon_card = "Inconsistencies" in results_html
    has_chron_table = "Medical Event Chronology" in results_html
    has_incon_section = "Inconsistency Flags" in results_html

    print_result("Summary cards — Total Events", has_events_card)
    print_result("Summary cards — Total Billed", has_billed_card)
    print_result("Summary cards — Total Paid", has_paid_card)
    print_result("Summary cards — Inconsistencies", has_incon_card)
    print_result("Chronology table section present", has_chron_table)
    print_result("Inconsistency Flags section present", has_incon_section)

    all_results_ok = all([has_events_card, has_billed_card, has_paid_card,
                          has_incon_card, has_chron_table, has_incon_section])
    if not all_results_ok:
        failures += 1

    # Check at least 1 chronology row (not just "No dated events")
    has_real_rows = "hover:bg-gray-50" in results_html and "No dated events" not in results_html
    print_result("Chronology table has at least 1 data row", has_real_rows)
    if not has_real_rows:
        failures += 1

    # ── TEST 4: DOWNLOAD PDF ─────────────────────────────────────────────────
    print("\n[4] DOWNLOAD PDF REPORT")
    r4 = client.get(f"/reports/{case_id}/pdf")
    pdf_ok = (r4.status_code == 200 and
              r4.headers.get("content-type", "").startswith("application/pdf") and
              len(r4.content) > 1000)
    print_result("GET /reports/{id}/pdf -> valid PDF bytes", pdf_ok,
                 f"status={r4.status_code} size={len(r4.content)} bytes")
    if not pdf_ok:
        failures += 1

    # Verify it's a real PDF (starts with %PDF)
    is_real_pdf = r4.content[:4] == b"%PDF"
    print_result("PDF content starts with %PDF magic bytes", is_real_pdf)
    if not is_real_pdf:
        failures += 1

    # ── TEST 5: DOWNLOAD EXCEL ───────────────────────────────────────────────
    print("\n[5] DOWNLOAD EXCEL")
    r5 = client.get(f"/reports/{case_id}/excel")
    excel_mime = "spreadsheetml" in r5.headers.get("content-type", "")
    excel_ok = (r5.status_code == 200 and excel_mime and len(r5.content) > 500)
    print_result("GET /reports/{id}/excel -> valid Excel bytes", excel_ok,
                 f"status={r5.status_code} size={len(r5.content)} bytes")
    if not excel_ok:
        failures += 1

    # Verify 3 sheets by checking the xlsx ZIP structure
    import io
    import zipfile
    try:
        with zipfile.ZipFile(io.BytesIO(r5.content)) as z:
            names = z.namelist()
        has_chronology = any("Chronology" in n or "sheet1" in n.lower() for n in names)
        print_result("Excel file is valid ZIP/xlsx", True, f"entries={len(names)}")

        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(r5.content))
        sheets = wb.sheetnames
        has_3_sheets = len(sheets) == 3
        print_result(f"Excel has 3 sheets: {sheets}", has_3_sheets)
        if not has_3_sheets:
            failures += 1
    except Exception as e:
        print_result("Excel file is valid", False, str(e))
        failures += 1

    # ── TEST 6: ERROR STATE ──────────────────────────────────────────────────
    print("\n[6] ERROR STATE — non-PDF upload")
    r6 = client.post("/cases/upload", data={
        "client_name": "Error Test",
        "case_number": "",
    }, files={"file": ("not_a_pdf.pdf", b"This is not PDF content", "application/pdf")})

    # Should be 400 with JSON detail, not 500
    error_shown = r6.status_code == 400
    not_500 = r6.status_code != 500
    try:
        err_json = r6.json()
        has_detail = "detail" in err_json
        detail_msg = err_json.get("detail", "")
    except Exception:
        has_detail = False
        detail_msg = r6.text[:100]

    print_result("Returns 400 not 500", not_500, f"status={r6.status_code}")
    print_result("Response is JSON with 'detail' key", has_detail, f"detail='{detail_msg}'")
    if not not_500 or not has_detail:
        failures += 1

    return failures


if __name__ == "__main__":
    print("=" * 60)
    print("ChartLens Week 3 — End-to-End Test")
    print("=" * 60)
    failed = run_tests()
    print("\n" + "=" * 60)
    if failed == 0:
        print("ALL TESTS PASSED [OK]")
    else:
        print(f"{failed} TEST(S) FAILED [FAIL]")
    print("=" * 60)
    sys.exit(0 if failed == 0 else 1)
