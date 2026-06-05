"""
Report generator: produces a PDF report and an Excel chronology from pipeline output.
No DB touches. No Claude API calls.
"""

import logging
from pathlib import Path

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, KeepTogether,
)
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_CHRON_COL_WIDTHS = [0.85 * inch, 1.5 * inch, 0.9 * inch, 2.10 * inch, 0.75 * inch, 0.9 * inch]


def _make_styles() -> dict:
    """Return a dict of named ParagraphStyle objects used throughout the PDF."""
    base = getSampleStyleSheet()
    return {
        "h1": ParagraphStyle("ReportH1", parent=base["Heading1"], fontSize=16, spaceAfter=8),
        "h2": ParagraphStyle("ReportH2", parent=base["Heading2"], fontSize=13, spaceAfter=6),
        "normal": base["Normal"],
        "small": ParagraphStyle("ReportSmall", parent=base["Normal"], fontSize=8, leading=10),
        "label": ParagraphStyle(
            "ReportLabel", parent=base["Normal"], fontSize=10, fontName="Helvetica-Bold"
        ),
    }


def _pdf_header_flowables(
    case_data: dict, chronology: dict, inconsistencies: dict, styles: dict
) -> list:
    """Return story elements for the case-summary header page."""
    case_number = case_data.get("case_number") or "N/A"
    client_name = case_data.get("client_name") or "Unknown"
    date_start = chronology.get("date_range_start") or "N/A"
    date_end = chronology.get("date_range_end") or "N/A"
    total_billed = float(chronology.get("total_billed") or 0.0)
    total_paid = float(chronology.get("total_paid") or 0.0)
    overall_risk = (inconsistencies.get("overall_risk_assessment") or "unknown").upper()
    total_incons = inconsistencies.get("total_inconsistencies") or 0
    high_sev = inconsistencies.get("high_severity_count") or 0
    total_events = chronology.get("total_events") or 0

    data = [
        ["Case Number:", case_number, "Client Name:", client_name],
        ["Date Range:", f"{date_start} to {date_end}", "Risk Assessment:", overall_risk],
        ["Total Billed:", f"${total_billed:,.2f}", "Total Paid:", f"${total_paid:,.2f}"],
        ["Total Events:", str(total_events), "Inconsistencies:", f"{total_incons} ({high_sev} high)"],
    ]
    tbl = Table(data, colWidths=[1.3 * inch, 2.2 * inch, 1.3 * inch, 2.2 * inch])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
    ]))
    return [
        Paragraph("ChartLens - Medical Records Analysis Report", styles["h1"]),
        Spacer(1, 0.15 * inch),
        tbl,
        PageBreak(),
    ]


def _pdf_chronology_flowables(chronology: dict, styles: dict) -> list:
    """Return story elements for the chronology table section."""
    flowables: list = [
        Paragraph("Medical Events Chronology", styles["h2"]),
        Spacer(1, 0.1 * inch),
    ]
    chron_list: list[dict] = chronology.get("chronology") or []

    if not chron_list:
        flowables.append(Paragraph("No dated events found in chronology.", styles["normal"]))
        return flowables

    headers = ["Date", "Provider", "Visit Type", "Diagnosis", "Billed", "Confidence"]
    table_data: list = [headers]
    for event in chron_list:
        diag_str = ", ".join(event.get("diagnosis") or []) or "-"
        billed = float(event.get("billed_amount") or 0.0)
        table_data.append([
            event.get("date") or "-",
            event.get("provider_name") or "-",
            event.get("visit_type") or "-",
            Paragraph(diag_str, styles["small"]),
            f"${billed:,.2f}",
            (event.get("confidence") or "-").capitalize(),
        ])

    tbl = Table(table_data, colWidths=_CHRON_COL_WIDTHS, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    flowables.append(tbl)

    undated_count = chronology.get("undated_count") or 0
    if undated_count:
        flowables.extend([
            Spacer(1, 0.1 * inch),
            Paragraph(
                f"Note: {undated_count} event(s) had no parseable date and are excluded above.",
                styles["normal"],
            ),
        ])
    return flowables


def _pdf_inconsistencies_flowables(inconsistencies: dict, styles: dict) -> list:
    """Return story elements for the inconsistencies section (forced to a new page)."""
    flowables: list = [
        PageBreak(),
        Paragraph("Inconsistency Flags", styles["h2"]),
        Spacer(1, 0.1 * inch),
    ]
    incons_list: list[dict] = inconsistencies.get("inconsistencies") or []

    if not incons_list:
        flowables.append(Paragraph("No inconsistencies detected.", styles["normal"]))
        return flowables

    _sev_labels = {"high": "HIGH", "medium": "MEDIUM", "low": "LOW"}
    for idx, item in enumerate(incons_list, start=1):
        inc_type = item.get("inconsistency_type") or item.get("type") or "Unknown"
        severity = item.get("severity") or "unknown"
        desc = item.get("description") or "No description provided."
        rec = item.get("recommendation") or "No recommendation."
        date1 = item.get("date_1") or "-"
        date2 = item.get("date_2") or "-"
        pages_str = ", ".join(str(p) for p in (item.get("page_references") or [])) or "-"
        label = _sev_labels.get(severity, severity.upper())
        flowables.append(KeepTogether([
            Paragraph(f"{idx}. [{label}] {inc_type}", styles["label"]),
            Paragraph(f"Dates: {date1} / {date2}  |  Pages: {pages_str}", styles["small"]),
            Paragraph(desc, styles["normal"]),
            Paragraph(f"<b>Recommendation:</b> {rec}", styles["normal"]),
            Spacer(1, 0.12 * inch),
        ]))
    return flowables


def generate_pdf_report(
    case_data: dict,
    chronology: dict,
    inconsistencies: dict,
    output_path: str,
) -> str | None:
    """
    Build a court-ready PDF report from pipeline output.

    Page 1 is a case summary; subsequent pages hold the chronology table;
    the final page lists inconsistency flags.
    Returns output_path on success, None on any failure.
    """
    try:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            leftMargin=0.75 * inch,
            rightMargin=0.75 * inch,
            topMargin=0.75 * inch,
            bottomMargin=0.75 * inch,
        )
        styles = _make_styles()
        story = (
            _pdf_header_flowables(case_data, chronology, inconsistencies, styles)
            + _pdf_chronology_flowables(chronology, styles)
            + _pdf_inconsistencies_flowables(inconsistencies, styles)
        )
        doc.build(story)
        logger.info("PDF report written to %s", output_path)
        return output_path
    except Exception as exc:
        logger.exception("generate_pdf_report failed: %s", exc)
        return None


def _autofit_columns(ws) -> None:
    """Set each column width to the widest cell value length, capped at 52 chars."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 52)


def _fill_chronology_sheet(ws, chronology: dict) -> None:
    """Write bold headers and one row per dated event to the Chronology sheet."""
    headers = [
        "Date", "Provider Name", "Provider Type", "Visit Type",
        "Diagnosis", "ICD Codes", "Treatment",
        "Billed Amount", "Paid Amount", "Page Numbers",
        "Confidence", "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for event in (chronology.get("chronology") or []):
        ws.append([
            event.get("date") or "",
            event.get("provider_name") or "",
            event.get("provider_type") or "",
            event.get("visit_type") or "",
            ", ".join(event.get("diagnosis") or []),
            ", ".join(str(c) for c in (event.get("icd_codes") or [])),
            event.get("treatment") or "",
            float(event.get("billed_amount") or 0.0),
            float(event.get("paid_amount") or 0.0),
            ", ".join(str(p) for p in (event.get("page_numbers") or [])),
            event.get("confidence") or "",
            event.get("notes") or "",
        ])
    _autofit_columns(ws)


def _fill_inconsistencies_sheet(ws, inconsistencies: dict) -> None:
    """Write bold headers and one row per inconsistency to the Inconsistencies sheet."""
    headers = [
        "Type", "Severity", "Description",
        "Date 1", "Date 2", "Page References", "Recommendation",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for item in (inconsistencies.get("inconsistencies") or []):
        ws.append([
            item.get("inconsistency_type") or item.get("type") or "",
            item.get("severity") or "",
            item.get("description") or "",
            item.get("date_1") or "",
            item.get("date_2") or "",
            ", ".join(str(p) for p in (item.get("page_references") or [])),
            item.get("recommendation") or "",
        ])
    _autofit_columns(ws)


def _fill_summary_sheet(
    ws, case_data: dict, chronology: dict, inconsistencies: dict
) -> None:
    """Write bold headers and key-value summary pairs to the Summary sheet."""
    ws.append(["Field", "Value"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in [
        ("Case Number", case_data.get("case_number") or ""),
        ("Client Name", case_data.get("client_name") or ""),
        ("Total Events", chronology.get("total_events") or 0),
        ("Undated Events", chronology.get("undated_count") or 0),
        ("Total Billed", chronology.get("total_billed") or 0.0),
        ("Total Paid", chronology.get("total_paid") or 0.0),
        ("Date Range Start", chronology.get("date_range_start") or ""),
        ("Date Range End", chronology.get("date_range_end") or ""),
        ("Overall Risk", inconsistencies.get("overall_risk_assessment") or ""),
        ("Total Inconsistencies", inconsistencies.get("total_inconsistencies") or 0),
        ("High Severity Count", inconsistencies.get("high_severity_count") or 0),
    ]:
        ws.append(list(row))
    _autofit_columns(ws)


def generate_excel_report(
    case_data: dict,
    chronology: dict,
    inconsistencies: dict,
    output_path: str,
) -> str | None:
    """
    Build a three-sheet Excel workbook from pipeline output.

    Sheets: Chronology (all events), Inconsistencies, Summary (key-value pairs).
    Returns output_path on success, None on any failure.
    """
    try:
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        ws_chron = wb.active
        ws_chron.title = "Chronology"
        _fill_chronology_sheet(ws_chron, chronology)

        ws_inc = wb.create_sheet("Inconsistencies")
        _fill_inconsistencies_sheet(ws_inc, inconsistencies)

        ws_sum = wb.create_sheet("Summary")
        _fill_summary_sheet(ws_sum, case_data, chronology, inconsistencies)

        wb.save(output_path)
        logger.info("Excel report written to %s", output_path)
        return output_path
    except Exception as exc:
        logger.exception("generate_excel_report failed: %s", exc)
        return None
